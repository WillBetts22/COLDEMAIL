"""
Spreadsheet output stage.
Writes one .xlsx file per run with one row per company.
Columns: Decision (dropdown), Company, Website, Contact Name, Title, Email,
         Fit Summary, Email Subject, Email Body, Sources, Apollo / Data Notes.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.config import AppConfig
from src.models import DraftResult

logger = logging.getLogger("coldemail")

# Header styling
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

# (column header, width in chars)
COLUMNS = [
    ("Decision", 13),
    ("Company", 26),
    ("Website", 30),
    ("Contact Name", 22),
    ("Title", 26),
    ("Email", 32),
    ("Fit Summary", 46),
    ("Email Subject", 36),
    ("Email Body", 58),
    ("Sources", 52),
    ("Apollo / Data Notes", 36),
]

# Columns that need word-wrap (0-indexed positions)
_WRAP_COLS = {6, 8, 9, 10}  # Fit Summary, Email Body, Sources, Notes


def write_spreadsheet(
    results: List[DraftResult],
    config: AppConfig,
    dry_run: bool = False,
) -> str:
    """Write all draft results to a timestamped .xlsx file. Returns the file path."""
    output_dir = Path(config.run.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    prefix = "DRY_RUN_" if dry_run else "outreach_batch_"
    filepath = output_dir / f"{prefix}{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outreach Batch"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _TOP
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Decision column dropdown (Approve / Skip)
    dv = DataValidation(type="list", formula1='"Approve,Skip"', allow_blank=True)
    dv.sqref = f"A2:A{max(len(results) + 1, 2)}"
    ws.add_data_validation(dv)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, result in enumerate(results, start=2):
        candidate = result.candidate
        company = candidate.company
        contact = candidate.contact
        research = result.research

        # Sources: one line per item, prefixed with source type
        source_lines: List[str] = []
        for item in research.research_items:
            label = {"website": "Website", "web_search": "Web search", "apollo": "Apollo"}.get(
                item.source_type, item.source_type
            )
            source_lines.append(f"{label}: {item.source_url}")
        sources_text = "\n".join(source_lines)

        # Notes: structured data + any warnings
        note_parts: List[str] = []
        if company.employee_count:
            note_parts.append(f"Employees: {company.employee_count:,}")
        if company.industry:
            note_parts.append(f"Industry: {company.industry}")
        if company.location:
            note_parts.append(f"Location: {company.location}")
        for note in research.notes:
            note_parts.append(f"⚠ {note}")
        if result.draft_failed:
            err = result.draft_error or "review manually"
            note_parts.append(f"⚠ DRAFT FAILED — {err}")
        notes_text = "\n".join(note_parts)

        subject = result.email.subject if result.email else "DRAFT FAILED — review manually"
        body = result.email.body if result.email else ""
        fit_summary = result.fit_summary or ("DRAFT FAILED — review manually" if result.draft_failed else "")

        row_data = [
            "",                                        # Decision — blank for user
            company.name,
            company.website or company.domain or "",
            contact.full_name,
            contact.title,
            contact.email,
            fit_summary,
            subject,
            body,
            sources_text,
            notes_text,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _WRAP_TOP if (col_idx - 1) in _WRAP_COLS else _TOP

        ws.row_dimensions[row_idx].height = 90

    wb.save(filepath)
    logger.info(f"Spreadsheet saved: {filepath}")
    return str(filepath)
